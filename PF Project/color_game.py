import tkinter as tk
from tkinter import messagebox
import random
import openpyxl
from openpyxl import Workbook
import os



#reference code: https://youtu.be/3WCkMEPN_uQ
# List of colors
colors = ["Red", "Blue", "Green", "Pink", "Black", "Yellow", "Orange", "White", "Purple", "Brown", "Cyan"]
score = 0
timeleft = 60

# Initialization
def startGame(event):
    # Start the game on Enter key press
    try:
        if timeleft == 60:
            countdown()  # Start the countdown timer
        nextcolor()  # Display the next color
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while starting the game: {e}")

def countdown():
    # Update the countdown
    global timeleft
    if timeleft == 0:
        messagebox.showinfo("Time Left", "Time is over and Your score is: " + str(score))
        store_data(score)
        show_highest_score()
    if timeleft > 0:
        timeleft -= 1
        timeLabel.config(text="Time Left :" + str(timeleft))
        timeLabel.after(1000, countdown)

def nextcolor():
    # Display the next color and update score
    global score
    global timeleft

    if timeleft > 0:
        e.focus_set()
        if e.get().lower() == colors[1].lower():
            score += 1

        e.delete(0, tk.END)
        random.shuffle(colors)

        label.config(fg=str(colors[1]), text=str(colors[0]))
        scoreLabel.config(text="Score: " + str(score))

def store_data(score):
    file_path = 'user_data.xlsx'
    
    try:
        if not os.path.exists(file_path):
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = 'User Data'
            sheet.append(["Score"])  # Add headers
            workbook.save(file_path)
        
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        sheet.append([score])
        
        workbook.save(file_path)
        
        print("Data stored successfully.")
    except Exception as e:
        print(f"Error storing data: {e}")

def show_highest_score():
    file_path = 'user_data.xlsx'
    
    try:
        if os.path.exists(file_path):
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            
            scores = []
            for row in range(2, sheet.max_row + 1, 2):  # Read only even rows
                score_value = sheet.cell(row=row, column=1).value
                if score_value is not None:
                    scores.append(score_value)
            
            if scores:
                highest_score = max(scores)
                messagebox.showinfo("Highest Score", f"The highest score is: {highest_score}")
            else:
                messagebox.showinfo("Highest Score", "No scores found.")
        else:
            messagebox.showinfo("Highest Score", "No data file found.")
    except Exception as e:
        print(f"Error reading data: {e}")

def open_game():
    # Set up and open the game window.
    global score
    global timeleft
    global timeLabel
    global scoreLabel
    global label
    global e

    # Reset score and time left
    score = 0
    timeleft = 60

    # Create a new game window    
    game_window = tk.Tk()
    game_window.title("My Color Game")
    game_window.geometry("400x650+400+100")
    game_window.resizable(0, 0)

    # Instructions label
    instructions = tk.Label(game_window, text="Type the color of the words, and not the word text!", font=('Helvetica', 12))
    instructions.pack(pady=20)

    # Score label
    scoreLabel = tk.Label(game_window, text="Press Enter to start", font=('Helvetica', 12))
    scoreLabel.pack(pady=10)
    
    # Time left label
    timeLabel = tk.Label(game_window, text="Time Left: " + str(timeleft), font=('Helvetica', 12))
    timeLabel.pack(pady=10)

    label = tk.Label(game_window, font=('Helvetica', 60))
    label.pack(pady=40)

    e = tk.Entry(game_window, font=('Helvetica', 24), justify='center')
    e.pack(pady=20, ipadx=10, ipady=5)
    game_window.bind("<Return>", startGame)
    e.focus_set()

